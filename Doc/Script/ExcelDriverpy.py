from TCParser import TCParser
import site
site.addsitedir(Project.Path+"TestRepo\\PyPackages")
import io

import openpyxl
from openpyxl import load_workbook
import numpy as np
import pandas as pd

class ExcelDriver:
    excelPath = ""
    workBook = ""
    
    def __init__(self,excelLoc):
        self.excelPath = excelLoc
        #in_mem_file=None
        
    def wbOpen(self,sheetName):
        Log.Message("Loading Worksheet")
        return self.workBook[sheetName]
        
    def loadWorkSheetIntoMem(self,inputSheetName):
        Log.Message("Trying to load worksheet into memory")
        df = pd.read_excel(self.excelPath, sheetname=inputSheetName,na_filter=False)
        #df1 = df.replace(np.nan, 'NAN', regex=True)
        dfToList = df.values.tolist()
        sheetHeader = list(df.columns)
        wsRec = [sheetHeader] + dfToList
        return wsRec
    
    
    def wbCreateSheet(self,sheetName):
      self.workBook = load_workbook(self.excelPath)
      self.workBook.create_sheet(sheetName)
      self.wbSave()
      
    # Appends data to existing sheet. Accepts list as data to write
    def wbAppend(self,datatoWrite,sheetName):
      self.workBook = load_workbook(self.excelPath)
      if sheetName in self.workBook.sheetnames:
        wSheet = self.workBook.get_sheet_by_name(sheetName)
      else: 
        wSheet = self.workBook.create_sheet(sheetName)
      for x in datatoWrite:
        wSheet.append(y.value for y in x)
      return self.workBook
    
    def wbWrite(self,datatoWrite,sheetName='Sheet1'):
      from pathlib import Path
      excelFile = Path(self.excelPath)
      if excelFile.exists():
        self.workBook = load_workbook(self.excelPath)
      else:
        from openpyxl import Workbook
        self.workBook = Workbook()
        
      if(sheetName=='Sheet1'):
        if sheetName in self.workBook.sheetnames:
          self.workBook.remove_sheet(self.workBook.get_sheet_by_name(sheetName)) 
        wSheet = self.workBook.create_sheet(sheetName)
      else:
        wSheet = self.workBook.create_sheet(sheetName)
        
      for x in datatoWrite:
        wSheet.append(y.value for y in x)
      return self.workBook  
    
    # close workbook after use without save
    def wbClose(self):
      self.workBook._archive.close()
      
    # Save Workbook to current path or specified path  
    def wbSave(self,Path=None):
      if Path is None:
        self.workBook.save(self.excelPath)
      else:
        self.workBook.save(Path)
        
    def getSheetNameList(self):
      self.workBook = load_workbook(self.excelPath)
      return self.workBook.sheetnames

#========================================== OLD CODE TO REMOVE AFTER TESTING ========================================================
        
    #Loads Excel Data into Memory
    def loadWorkSheetIntoMem_OpenPyxl(self,sheetName,isTCSheet=False):
        Log.Message("Trying to load worksheet into memory")
        self.workBook = load_workbook(self.excelPath, read_only=True, data_only=True)
        woSheet = self.wbOpen(sheetName)
        
        shHeader = {}
        wsRec=[]
        
        #Read header and decide what are column names and their positions
        shHeadRec = list(woSheet.rows)[0]
        for colHead in shHeadRec:
          if colHead.value is not None:
            shHeader[VarToString(colHead.value)]=colHead.column
          
        if(isTCSheet):
          colNumExecCond = shHeader['ExecCond']
          #Load All Excel Data into where exec condition is Y 
          for row in woSheet.rows:
            if row[colNumExecCond-1].value in ('Y','y','ExecCond'):
              wsRec.append([cell.value if cell.value is not None else "NONE" for cell in row])
        else:
          for row in woSheet.rows:
            wsRec.append([cell.value if cell.value is not None else "NONE" for cell in row])
        return wsRec